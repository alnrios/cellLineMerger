import openpyxl
import os
from classes import ExcelWkbk

directory = os.listdir('spreadsheets')


def create_excel_file():
    master_wkbk = openpyxl.Workbook()
    sheet = master_wkbk.active
    sheet['A1'].value = 'TreatmentBarcode'
    sheet['B1'].value = 'Treatment position'
    sheet['C1'].value = 'Staff_ID'
    sheet['D1'].value = 'Cell_Line_ID'
    sheet['E1'].value = 'SetupDate'
    sheet['F1'].value = 'Drug_ID_1'
    sheet['G1'].value = 'Starting_Concentration_in_uM'
    sheet['H1'].value = 'Dilution_Factor'
    sheet['I1'].value = 'Day1Barcode'
    sheet['J1'].value = 'Day1Location'
    sheet['K1'].value = 'Day7Barcode'
    sheet['L1'].value = 'Day7Location'
    master_sheet_name = input("Enter name of master sheet (DO NOT INCLUDE FILE EXTENSION): ")
    master_wkbk.save('{}.xlsx'.format(master_sheet_name))
    return master_sheet_name


def add_to_master(master_file, excelwkbk, row):
    master_wb = openpyxl.load_workbook(master_file)
    write_sheet = master_wb.active
    for sheet in excelwkbk.sheets:
        for x in range(len(sheet.treatment_list)):
            for cell_line in range(sheet.treatment_list[x].num_of_cell_lines):
                col = 1
                write_sheet.cell(column=col, row=row, value="{0}".format(sheet.treatment_list[x].barcode))
                col += 1
                write_sheet.cell(column=col, row=row,
                                 value="{0}".format(sheet.treatment_list[x].cell_line_list[cell_line].position))
                col += 1
                write_sheet.cell(column=col, row=row, value="{0}".format(excelwkbk.technician))
                col += 1
                current_cell_line = sheet.treatment_list[x].cell_line_list[cell_line]
                write_sheet.cell(column=col, row=row,
                                 value="{0}".format(current_cell_line))
                col += 1
                write_sheet.cell(column=col, row=row, value="{0}".format(sheet.date))
                col += 1
                write_sheet.cell(column=col, row=row, value="{0}".format(sheet.treatment_list[x].drug))
                col += 1
                write_sheet.cell(column=col, row=row, value=sheet.treatment_list[x].drug.concentration)
                col += 1
                write_sheet.cell(column=col, row=row, value=sheet.treatment_list[x].drug.dilution)
                col += 1
                try:
                    write_sheet.cell(column=col, row=row, value=sheet.treatment_list[x].day1_link.barcode)
                    col += 1
                    for d1 in range(sheet.treatment_list[x].day1_link.num_of_cell_lines):
                        if current_cell_line.name == sheet.treatment_list[x].day1_link.cell_line_list[d1].name:
                            write_sheet.cell(column=col, row=row,
                                             value=sheet.treatment_list[x].day1_link.cell_line_list[d1].position)
                            break
                        else:
                            write_sheet.cell(column=col, row=row, value='NA')
                    col += 1
                    write_sheet.cell(column=col, row=row, value=sheet.treatment_list[x].day7_link.barcode)
                    col += 1
                    for d7 in range(sheet.treatment_list[x].day7_link.num_of_cell_lines):
                        if current_cell_line.name == sheet.treatment_list[x].day7_link.cell_line_list[d7].name:
                            write_sheet.cell(column=col, row=row,
                                             value=sheet.treatment_list[x].day7_link.cell_line_list[d7].position)
                            break
                        else:
                            write_sheet.cell(column=col, row=row, value='NA')
                except AttributeError:
                    os.remove('{}.xlsx'.format(master))
                    raise AttributeError("Something is wrong with the cell lines input in this workbook: "
                                         + excelwkbk.technician + " - " + sheet.date + " - " +
                                         sheet.treatment_list[x].cell_line_list[cell_line].name +
                                         "\nFix the mentioned spreadsheet before continuing...")
                row += 1
    master_wb.save('{}.xlsx'.format(master))
    return row


master = create_excel_file()
row = 2
for i in range(len(directory)):
    wb = openpyxl.load_workbook('spreadsheets/{}'.format(directory[i]), data_only=True)
    new_wb = ExcelWkbk(wb)
    new_wb.get_worksheets()
    new_wb.display_sheet_info()
    print("\nWriting to master file...\n")
    row = add_to_master('{}.xlsx'.format(master), new_wb, row)
print("Script finished successfully!")
