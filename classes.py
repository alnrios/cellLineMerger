import openpyxl


def cell_scan(wkst, desired_value, col_num, type_of_scan):
    return_list = []  # used in the case that this function needs to return a list
    for col in wkst.iter_cols(col_num):
        for cell in col:
            if cell.value is not None:
                if type_of_scan == 'value return bool':
                    if desired_value in str(cell.value):
                        return True
                elif type_of_scan == 'value return value':
                    if desired_value in str(cell.value):
                        return cell
                elif type_of_scan == 'cell return cell':
                    if desired_value in str(cell):
                        return cell
                elif type_of_scan == 'cell return bool':
                    if desired_value in str(cell):
                        return True
                elif type_of_scan == 'value return list':  # returns list of cells that the value was found at
                    if desired_value in str(cell.value):
                        return_list.append(cell)
                else:
                    raise ValueError("Incorrect type of scan. Exit and try again...")
    if type_of_scan == 'value return list':
        return return_list


def control_scan(wkst, desired_value, col_num):
    control_list = []
    for col in wkst.iter_cols(col_num):
        for cell in col:
            if cell.value is not None:
                if isinstance(desired_value, list):
                    for i in range(len(desired_value)):
                        if desired_value[i] in str(cell.value):
                            control_list.append(clean_cell_name(cell))
                else:
                    if desired_value in str(cell.value):
                        control_list.append(clean_cell_name(cell))
    return control_list


def inc_by_column(cell_name, quantity):
    return chr(ord(cell_name[0]) + quantity) + cell_name[1:]


def inc_by_row(cell_name, quantity):
    return cell_name[0] + str(int(cell_name[1:]) + quantity)


def string_extract(cell, split_char):
    cell_string = cell.value
    if split_char in cell_string:
        return_val = cell_string.split(split_char)
        return "".join(return_val).strip()
    else:
        print(split_char + " not found in cell A1\n")


def clean_cell_name(cell):
    return str(cell).split('.')[1].split('>')[0]


def drug_check(wkst, col_number):
    for j in reversed(range(10)):
        if cell_scan(wkst, "Drug {}".format(str(j + 1)), col_number, "value return bool"):
            return j + 1


def get_control_cell_lines(wkst, barcode_cell, count):
    cell_line_list = []
    pos = 0  # records the position of the cell line
    for i in range(8):
        pos += 1
        barcode_cell = inc_by_column(barcode_cell, 1)
        count += 1
        cell_line = cell_scan(wkst, barcode_cell, count, 'cell return cell')
        if cell_line is not None and 'Cell Line' not in str(cell_line.value):
            cell_line_list.append(CellLine(str(cell_line.value), pos))
    return cell_line_list


class ExcelWkbk:
    def __init__(self, wkbk):
        self.imported_wkbk = wkbk
        self.sheet_names = wkbk.sheetnames
        self.sheets = []
        self.technician = ""

    def get_worksheets(self):
        ws = self.imported_wkbk[self.sheet_names[0]]
        self.technician = string_extract(ws['A1'], 'Name:')
        for sheet in range(len(self.sheet_names)):
            ws = self.imported_wkbk[self.sheet_names[sheet]]
            self.sheets.append(
                ExcelSheet(
                    date=string_extract(ws['A2'], 'SET UP DATE:'),
                    num_of_drugs=drug_check(ws, 1)
                )
            )
            num_drugs = self.sheets[sheet].num_of_drugs
            self.get_and_set_drugs(num_drugs, ws, sheet)
            self.get_controls(ws, sheet)
            self.get_treatments(ws, sheet)

    def get_and_set_drugs(self, num_drugs, wkst, sheet_index):
        for i in range(num_drugs):
            if cell_scan(wkst, 'Drug {}'.format(i + 1), 1, 'value return bool'):
                cell_val = cell_scan(wkst, 'Drug {}'.format(i + 1), 1, 'value return value')
                drug_name = cell_val.value
                cell_val = clean_cell_name(cell_val)
                conc_cell = inc_by_column(cell_val, 1)
                dilut_cell = inc_by_column(cell_val, 2)
                for col in wkst.iter_cols(2, 3):
                    for cell in col:
                        if conc_cell in str(cell):
                            conc_cell = cell.value
                            break
                        elif dilut_cell in str(cell):
                            dilut_cell = cell.value
                            break
                self.sheets[sheet_index].append_to_drug_list(Drug(drug_name, conc_cell, dilut_cell))

    def get_num_of_sheets(self):
        return len(self.sheets)

    def get_controls(self, wkst, sheet_index):
        control_xl_dict = {
            'd1_xl_cells': control_scan(wkst, ['Day1', 'Day 1'], 1),
            'd7_xl_cells': control_scan(wkst, ['Day7', 'Day 7'], 1)
        }
        for i in control_xl_dict:
            for xl_iter in range(len(control_xl_dict[i])):  # iterator for the control_xl_dict items in lists
                barcode = inc_by_column(control_xl_dict[i][xl_iter], 1)  # first assigns barcode var to the cell number
                col_count = 2  # col_count refers to the column number that'll be used, which will always be 2 for bc's
                cell_line_list = get_control_cell_lines(wkst, barcode, col_count)
                barcode = cell_scan(wkst, barcode, col_count, 'cell return cell').value
                if i == 'd1_xl_cells':
                    self.sheets[sheet_index].day1_list.append(
                        ControlPlate(barcode, len(cell_line_list), cell_line_list, 'd1'))
                elif i == 'd7_xl_cells':
                    self.sheets[sheet_index].day7_list.append(
                        ControlPlate(barcode, len(cell_line_list), cell_line_list, 'd7'))

    def display_sheet_info(self):
        print('*' * 35)
        print(str(self.get_num_of_sheets()) + " sheets in " + self.technician + "'s workbook")
        print('*' * 35)
        for i in range(len(self.sheets)):
            self.sheets[i].display_info()

    def get_treatments(self, wkst, sheet_index):
        if cell_scan(wkst, 'TREATMENT PLATES', 1, 'value return bool'):
            # below is a list comprised of cells that contain "TREATMENT PLATES as its value"
            treat_cells = cell_scan(wkst, 'TREATMENT PLATES', 1, 'value return list')
            for i in range(len(treat_cells)):
                count = 0  # this count var is necessary to know how many cells to go backwards when capturing
                # cell line positions (ie. Column 1-2, Column 3-4)
                for j in range(self.sheets[sheet_index].num_of_drugs):
                    cleaned_cell = clean_cell_name(treat_cells[i])
                    cleaned_cell = inc_by_row(cleaned_cell, j + 1)
                    drug = wkst[cleaned_cell].value
                    cleaned_cell = inc_by_column(cleaned_cell, 1)
                    if wkst[cleaned_cell].value is not None and 'Treatment' not in wkst[cleaned_cell].value:
                        barcode = wkst[cleaned_cell].value
                        cell_lines = []
                        for k in range(4):
                            cleaned_cell = inc_by_column(cleaned_cell, 1)
                            position_cell = inc_by_row(cleaned_cell, count - 1)
                            if wkst[cleaned_cell].value is not None and 'Cell' not in str(wkst[cleaned_cell].value):
                                cell_lines.append(CellLine(str(wkst[cleaned_cell].value), wkst[position_cell].value))
                        for listed_drug in self.sheets[sheet_index].drug_list:
                            if drug == listed_drug.name:
                                drug = listed_drug
                        import_treatment = TreatmentPlate(barcode, len(cell_lines), cell_lines, drug)
                        import_treatment.get_control_links('d1', self.sheets[sheet_index])
                        import_treatment.get_control_links('d7', self.sheets[sheet_index])
                        self.sheets[sheet_index].treatment_list.append(import_treatment)
                    count -= 1
                    '''
                    ^^^^^^^^
                    The count variable is subtracted above here this far back in the loop because the positions will get
                    further from the actual cell line that is being processed. Simply subtracting 1 every time will not
                    cut it, especially when trying to account for multiple drugs that could reach a count of 5.
                    '''


class ExcelSheet:
    def __init__(self, date, num_of_drugs):
        self.date = date
        self.num_of_drugs = num_of_drugs
        self.drug_list = []
        self.day1_list = []
        self.day7_list = []
        self.treatment_list = []

    def display_info(self):
        print(str(self.num_of_drugs) + " drugs used on " + self.date)
        for i in range(self.num_of_drugs):
            drug_info = self.drug_list[i]
            print(drug_info.name + "\n\tConcentration: " +
                  str(drug_info.concentration) + "\n\tDilution: " + str(drug_info.dilution))
        print('-' * 35)
        print("Control Plates:")
        for i in range(len(self.day1_list)):
            print("\t" + str(self.day1_list[i]))
        for j in range(len(self.day7_list)):
            print("\t" + str(self.day7_list[j]))
        print("Treatment Plates")
        for k in range(len(self.treatment_list)):
            print("\t" + str(self.treatment_list[k]))

    def append_to_drug_list(self, drug):
        self.drug_list.append(drug)


class Plate:
    def __init__(self, barcode, num_of_cell_lines, cell_line_list):
        self.barcode = barcode
        self.num_of_cell_lines = num_of_cell_lines
        self.cell_line_list = cell_line_list

    def display_cell_lines(self):
        return_string = "["
        for i in range(self.num_of_cell_lines):
            if i == self.num_of_cell_lines - 1:
                return_string += (self.cell_line_list[i].name + ": " + str(self.cell_line_list[i].position))
            else:
                return_string += (self.cell_line_list[i].name + ": " + str(self.cell_line_list[i].position) + ", ")
        return_string += "]"
        return return_string


class ControlPlate(Plate):
    def __init__(self, barcode, num_of_cell_lines, cell_line_list, control_type):
        super().__init__(barcode, num_of_cell_lines, cell_line_list)
        self.control_type = control_type

    def __str__(self):
        return str(self.num_of_cell_lines) + " cell lines in " + self.control_type + " plate - " + \
               self.barcode + ": " + self.display_cell_lines()

    def display_cell_lines(self):
        return_string = "["
        for i in range(self.num_of_cell_lines):
            if i == self.num_of_cell_lines - 1:
                return_string += (self.cell_line_list[i].name + ": " + str(self.cell_line_list[i].position))
            else:
                return_string += (self.cell_line_list[i].name + ": " + str(self.cell_line_list[i].position) + ", ")
        return_string += "]"
        return return_string


class TreatmentPlate(Plate):
    def __init__(self, barcode, num_of_cell_lines, cell_line_list, drug):
        super().__init__(barcode, num_of_cell_lines, cell_line_list)
        self.drug = drug
        self.day1_link = ""
        self.day7_link = ""

    def get_control_links(self, control_type, wkst):
        if control_type == 'd1':
            control_list = wkst.day1_list
        elif control_type == 'd7':
            control_list = wkst.day7_list
        else:
            raise ValueError("Incorrect control type inputted into the 'get_control_links' function")
        for control_plate in control_list:
            pass_var = True
            ctrl_cell_line_list = []
            for i in range(len(control_plate.cell_line_list)):
                ctrl_cell_line_list.append(control_plate.cell_line_list[i].name)
            for cell_line in self.cell_line_list:
                if cell_line.name not in ctrl_cell_line_list:
                    pass_var = False
                    break
            if pass_var and control_type == 'd1':
                self.day1_link = control_plate
            elif pass_var and control_type == 'd7':
                self.day7_link = control_plate
            else:
                pass

    def __str__(self):
        return self.barcode + ": " + self.display_cell_lines() + " treated with " + str(self.drug)


class Drug:
    def __init__(self, name, concentration, dilution):
        self.name = name
        self.concentration = concentration
        self.dilution = dilution

    def __str__(self):
        return self.name


class CellLine:
    def __init__(self, name, position):
        self.name = name
        self.position = position

    def __str__(self):
        return self.name
